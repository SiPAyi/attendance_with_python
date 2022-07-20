from time import ctime, time
import getdetails
import perioddetails
import getattendence
import openpyxl as xl

details = getdetails.getDetails("attendence.xlsx")
period_name = perioddetails.periodName()
# period_name = 'maths'


if period_name not in ['break', 'no class']:
    wb = xl.load_workbook('attendence.xlsx')
    sheet = wb[period_name]
    last_col_number = sheet.max_column

def uploadDetails(sheet):
    student_state = getattendence.takeAttendence()
    for rollNo in student_state:
        sheet.cell(rollNo+1, last_col_number+1).value = student_state[rollNo]

def overWrite(sheet):
    student_state = getattendence.takeAttendence()
    for rollNo in student_state:
        sheet.cell(rollNo+1, last_col_number).value = student_state[rollNo]    

date = ctime(time())[8:10] + '-' + ctime(time())[4:7] + '-' +ctime(time())[20:25] # 20-Jul-2020
todays_col_first_row = sheet.cell(1, last_col_number+1)
# print(todays_col_first_row)



if sheet.cell(1, last_col_number).value != date:
    todays_col_first_row.value = date
    uploadDetails(sheet)
elif sheet.cell(1, last_col_number).value == 'ID':
    uploadDetails(sheet)
else:
    print('already entered today\'s attendence. do you want to overwrite them ?')    
    overwrite = input('Enter your opinion (y/n) : ')
    if overwrite == 'y':
        overWrite(sheet)
        print('overwrited')
    else:
        print('thanks :) ')

wb.save("attendence.xlsx")