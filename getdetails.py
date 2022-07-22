import openpyxl as xl

# to get details as a list like the below one
# ['roll.no', 'id number', 'name']

def getDetailsList(workbook):
    wb = xl.load_workbook(workbook)
    students_list_sheet = wb['student_list']

    students_list = []

    for row in students_list_sheet:
        student_details = [cell.value for cell in row]
        students_list.append(student_details)

    return students_list


# to get details in a dictionary format like the below one
# students = {
#     1 : {'id':'ro200174', 'name':"sai"}
# }

def getDetails(workbook):
    wb = xl.load_workbook(workbook)
    students_list_sheet = wb['student_list']

    students_list = {}

    for i in range(2, 74):
        rollNo = students_list_sheet.cell(i, 1).value
        id = students_list_sheet.cell(i, 2).value
        name = students_list_sheet.cell(i, 3).value
        students_list[rollNo] = {'id':id, 'name':name}

    return students_list
